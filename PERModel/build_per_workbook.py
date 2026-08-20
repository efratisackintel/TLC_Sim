"""Human-readable workbook of ALL PER permutations, from per_catalog.npz.

Permutation = Format x FrameBW x ChannelBW x Antenna x GI  (135 total), each with MCS 0..15.

Sheets:
  * Sensitivity SNR@10%  - wide: Format|FrameBW|ChannelBW|Ant|GI|Band|MCS0..MCS15 -> SNR(dB) at 10% PER
  * Peak Throughput Mbps - same wide layout -> ideal Mbps per rate
  * Index                - one row per curve
  * PER <frameBW> <ant> <gi>  - full-channel per-case sheets: SNR grid x MCS PER + a
                                log-Y scatter chart with all MCS overlaid.

Run from repo root (after build_per_catalog.py):
    python PERModel/build_per_workbook.py
"""
from __future__ import annotations

import os

import numpy as np
from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CAT = "PERModel/per_catalog.npz"
OUT = "PERModel/PER_Model_AllPermutations.xlsx"
GRID_STEP = 0.5
FLOOR = 1e-4                                  # log-axis floor for PER=0 points

BASE_TPT = [30.6, 61.2, 91.8, 122.5, 183.7, 245, 275.6, 306.2,
            367.5, 408.3, 459.3, 510.4, 547, 604]      # 80MHz SISO 0.8us, MCS 0..13
BW_FACTOR = {20: 0.25, 40: 0.5, 80: 1, 160: 2, 320: 4}
GI_F = {0.8: 1.0, 1.6: 0.944, 3.2: 0.85}     # symbol-time throughput factor
BWS = [20, 40, 80, 160, 320]
ANT_ORDER = ["1x1", "1x2", "2x2"]
GI_ORDER = [0.8, 1.6, 3.2]
MAX_MCS = 15


def _load():
    d = np.load(CAT)
    keys = [str(k) for k in d["keys"]]
    curve = {k: (d[f"x{i}"], d[f"y{i}"]) for i, k in enumerate(keys)}
    sens = {k: float(d["snr_sens"][i]) for i, k in enumerate(keys)}
    return curve, sens


def _nss(ant):
    return 2 if ant == "2x2" else 1


def _peak_tpt(fbw, ant, gi, mcs):
    if mcs >= len(BASE_TPT):
        return None
    return round(BASE_TPT[mcs] * BW_FACTOR[fbw] * _nss(ant) * GI_F[gi], 1)


def build():
    curve, sens = _load()
    have = lambda fbw, chbw, ant, gi: any(
        f"{fbw},{chbw},{ant},{gi},{m}" in curve for m in range(MAX_MCS + 1))
    perms = [(fbw, chbw, ant, gi)
             for fbw in BWS for chbw in BWS if chbw >= fbw
             for ant in ANT_ORDER for gi in GI_ORDER if have(fbw, chbw, ant, gi)]

    wb = Workbook()
    wb.remove(wb.active)
    hfill = PatternFill("solid", fgColor="1F2A44")
    hfont = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center")
    META = ["Format", "Frame BW", "Channel BW", "Ant", "GI (us)", "Band"]

    def _wide(title, cellfn, numfmt):
        ws = wb.create_sheet(title)
        heads = META + [f"MCS{m}" for m in range(MAX_MCS + 1)]
        for j, h in enumerate(heads, start=1):
            c = ws.cell(row=1, column=j, value=h)
            c.font = hfont; c.fill = hfill; c.alignment = center
        for r, (fbw, chbw, ant, gi) in enumerate(perms, start=2):
            for j, v in enumerate(["EHT", fbw, chbw, ant, gi, None], start=1):
                ws.cell(row=r, column=j, value=v)
            for m in range(MAX_MCS + 1):
                v = cellfn(fbw, chbw, ant, gi, m)
                cell = ws.cell(row=r, column=len(META) + 1 + m, value=v)
                if v is not None:
                    cell.number_format = numfmt; cell.alignment = center
        ws.freeze_panes = "G2"
        for col, w in zip("ABCDEF", (7, 9, 11, 6, 8, 6)):
            ws.column_dimensions[col].width = w
        for m in range(MAX_MCS + 1):
            ws.column_dimensions[get_column_letter(len(META) + 1 + m)].width = 7

    _wide("Sensitivity SNR@10%",
          lambda fbw, chbw, ant, gi, m: (
              round(sens[f"{fbw},{chbw},{ant},{gi},{m}"], 1)
              if f"{fbw},{chbw},{ant},{gi},{m}" in sens
              and not np.isnan(sens[f"{fbw},{chbw},{ant},{gi},{m}"]) else None),
          "0.0")
    _wide("Peak Throughput Mbps",
          lambda fbw, chbw, ant, gi, m: (
              _peak_tpt(fbw, ant, gi, m)
              if f"{fbw},{chbw},{ant},{gi},{m}" in curve else None),
          "0.0")

    # Index
    idx = wb.create_sheet("Index")
    heads = ["Format", "Frame BW", "Channel BW", "Ant", "GI (us)", "MCS",
             "Sensitivity SNR@10% (dB)", "Points", "SNR start", "SNR end", "Peak TpT (Mbps)"]
    for j, h in enumerate(heads, start=1):
        c = idx.cell(row=1, column=j, value=h); c.font = hfont; c.fill = hfill
    r = 2
    for fbw, chbw, ant, gi in perms:
        for m in range(MAX_MCS + 1):
            k = f"{fbw},{chbw},{ant},{gi},{m}"
            if k not in curve:
                continue
            x, _ = curve[k]; s = sens.get(k)
            vals = ["EHT", fbw, chbw, ant, gi, m,
                    None if s is None or np.isnan(s) else round(s, 1),
                    int(x.size), round(float(x.min()), 1), round(float(x.max()), 1),
                    _peak_tpt(fbw, ant, gi, m)]
            for j, v in enumerate(vals, start=1):
                idx.cell(row=r, column=j, value=v)
            r += 1
    idx.freeze_panes = "A2"

    # Full-channel per-case PER sheets with log-Y all-MCS chart
    # Per-case PER sheets (ALL permutations) with log-Y all-MCS chart
    for fbw, chbw, ant, gi in perms:
        mcss = [m for m in range(MAX_MCS + 1)
                if f"{fbw},{chbw},{ant},{gi},{m}" in curve]
        if not mcss:
            continue
        lo = min(curve[f"{fbw},{chbw},{ant},{gi},{m}"][0].min() for m in mcss)
        hi = max(curve[f"{fbw},{chbw},{ant},{gi},{m}"][0].max() for m in mcss)
        lo = np.floor(lo / GRID_STEP) * GRID_STEP
        hi = np.ceil(hi / GRID_STEP) * GRID_STEP
        grid = np.round(np.arange(lo, hi + GRID_STEP / 2, GRID_STEP), 3)

        gtag = '08' if gi == 0.8 else '16' if gi == 1.6 else '32'
        name = f"PER {fbw}-{chbw} {ant} {gtag}us"
        ws = wb.create_sheet(name[:31])
        ws.cell(row=1, column=1, value="SNR (dB)").font = hfont
        ws.cell(row=1, column=1).fill = hfill
        for gi_i, g in enumerate(grid, start=2):
            ws.cell(row=gi_i, column=1, value=float(g))
        for j, m in enumerate(mcss, start=2):
            x, y = curve[f"{fbw},{chbw},{ant},{gi},{m}"]
            o = np.argsort(x)
            yg = np.interp(grid, x[o], y[o], left=1.0, right=0.0)
            yg = np.clip(yg, FLOOR, 1.0)              # log-safe
            c = ws.cell(row=1, column=j, value=f"MCS{m}")
            c.font = hfont; c.fill = hfill
            for gi_i, v in enumerate(yg, start=2):
                ws.cell(row=gi_i, column=j, value=round(float(v), 5))
            ws.column_dimensions[get_column_letter(j)].width = 8
        ws.column_dimensions["A"].width = 9
        ws.freeze_panes = "B2"

        chart = ScatterChart()
        chart.title = f"PER vs SNR - EHT {fbw}/{chbw} MHz {ant} {gi}us (all MCS)"
        chart.x_axis.title = "SNR (dB)"
        chart.y_axis.title = "PER"
        chart.y_axis.scaling.logBase = 10
        chart.y_axis.scaling.min = FLOOR
        chart.y_axis.scaling.max = 1.0
        chart.height = 11; chart.width = 22
        xref = Reference(ws, min_col=1, min_row=2, max_row=len(grid) + 1)
        for j, m in enumerate(mcss, start=2):
            yref = Reference(ws, min_col=j, min_row=1, max_row=len(grid) + 1)
            s = Series(yref, xref, title_from_data=True)
            s.marker.symbol = "none"
            s.graphicalProperties.line.width = 18000
            chart.series.append(s)
        ws.add_chart(chart, f"{get_column_letter(len(mcss) + 3)}2")

    wb.save(OUT)
    print(f"Saved {OUT}: {len(perms)} permutations, "
          f"{os.path.getsize(OUT)/1024:.0f} KB, {len(wb.sheetnames)} sheets")


if __name__ == "__main__":
    build()
