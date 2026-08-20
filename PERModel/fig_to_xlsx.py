"""Export raw curve data from MATLAB .fig files into an .xlsx workbook.

Each .fig stores many Line objects (one per test config). This writes,
per .fig:
  * a data sheet: a single shared "SNR (dB)" column on a uniform 0.5 dB grid,
    then one metric column per test. Each test is resampled onto the grid
    (linear interpolation inside its swept range); outside its range the value
    saturates to 1.0 (100% error) below and 0.0 (0% error) above.
  * a legend sheet: the config parsed into individual columns (Test, Format,
    SU/MU, FrameBW, ChannelBW, antenna, Nss, GI, LTF, coding, length, ...)
    plus SNR_sens and the sensitivity target.

Usage:
    python fig_to_xlsx.py OUT.xlsx FILE1.fig [FILE2.fig ...]
"""
from __future__ import annotations

import argparse
import os
import re

import numpy as np
from scipy.io import loadmat

from fig_viewer import _iter_children, _node, _prop, _flatten_str

ERROR_METRICS = {"BER", "PER", "DPER", "HPER"}
GRID_STEP = 0.5


def extract_curves(path: str):
    """Return list of dicts {name, x, y} for every line in the .fig."""
    mat = loadmat(path, squeeze_me=False, struct_as_record=True)
    root_key = next((k for k in mat if k.startswith("hgS")), None)
    if root_key is None:
        raise SystemExit(f"No hgS figure struct found in {path}")

    curves: list[dict] = []

    def walk(node):
        ctyp, cprops, children = _node(node)
        if ctyp in ("graph2d.lineseries", "line"):
            x = np.asarray(_prop(cprops, "XData"), dtype=float).ravel()
            y = np.asarray(_prop(cprops, "YData"), dtype=float).ravel()
            if x.size and y.size:
                name = _flatten_str(_prop(cprops, "DisplayName") or "")
                curves.append({"name": str(name), "x": x, "y": y})
        for c in _iter_children(children):
            walk(c)

    for n in np.asarray(mat[root_key]).ravel():
        walk(n)
    return curves


def clean_name(s: str) -> str:
    """Turn a MATLAB TeX DisplayName into readable plain text."""
    s = s.replace("$", "")
    s = s.replace("\\times", " x ").replace("\\&", "&")
    s = re.sub(r"\\\\", " ", s)          # TeX line break
    s = re.sub(r"\\[, ]", " ", s)         # thin space / escaped space
    s = s.replace("_{", "_").replace("^{", "^").replace("{", "").replace("}", "")
    s = re.sub(r"\\", "", s)              # any stray backslash
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"\s*,\s*", ", ", s)
    s = re.sub(r"(,\s*){2,}", ", ", s)    # collapse repeated commas
    return s.strip().strip(",").strip()


def _search(pat, s, grp=1, cast=str, default=None):
    m = re.search(pat, s)
    if not m:
        return default
    try:
        return cast(m.group(grp))
    except (ValueError, IndexError):
        return default


def parse_config(cleaned: str) -> dict:
    """Break a cleaned DisplayName into individual named fields."""
    fbw = re.search(r"(\d+)\s*/\s*(\d+)", cleaned)
    ant = re.search(r"(\d+)\s*x\s*(\d+)", cleaned)          # first "A x B" = antenna
    ltf = re.search(r"LTF[^,]*?x\s*(\d+)", cleaned)          # "...& x 2" -> 2
    tgt = re.search(r"([A-Za-z]+)_?sens\s*=\s*(10\^-?\d+)", cleaned)
    return {
        "Test #": _search(r"Test\s*(\d+)", cleaned, cast=int),
        "Format": _search(r"\b(EHT|HE|VHT|HT|UHR|Legacy|OFDM|DSSS|CCK)\b", cleaned),
        "SU/MU": _search(r"\b(SU|MU)\b", cleaned),
        "FrameBW": int(fbw.group(1)) if fbw else None,
        "ChannelBW": int(fbw.group(2)) if fbw else None,
        "Antenna": f"{ant.group(1)}x{ant.group(2)}" if ant else None,
        "Nss": _search(r"N\s*S\s*(\d+)", cleaned, cast=int),
        "GI [us]": _search(r"LTF\s*([\d.]+)", cleaned, cast=float),
        "LTF": int(ltf.group(1)) if ltf else None,
        "LDPC": _search(r"(?<![A-Za-z0-9])LDPC\s+(\d+)", cleaned, cast=int),
        "LDPC2xCW": _search(r"LDPC2xCW\s+(\d+)", cleaned, cast=int),
        "Length [B]": _search(r"Length\s*\[B\]\s*(\d+)", cleaned, cast=int),
        "PPM": _search(r"PPM\s+(-?\d+)", cleaned, cast=int),
        "Channel": _search(r"([A-Za-z0-9]+)\s*CH\b", cleaned),
        "Control": _search(r"(?<![A-Za-z])Control\s+([A-Za-z0-9]+)", cleaned),
        "RandomControl": _search(r"RandomControl\s+(\d+)", cleaned, cast=int),
        "LFD": _search(r"\bLFD\s+(\d+)", cleaned, cast=int),
        "Lean2Main": _search(r"Lean2Main\s+(\d+)", cleaned, cast=int),
        "RUsize": _search(r"RUsize\s+(\d+)", cleaned, cast=int),
        "RUindex": _search(r"RUindex\s+(\d+)", cleaned, cast=int),
        "SC": _search(r"SC\s*(\[\]|\S+)", cleaned),
        "SNR_sens [dB]": _search(r"SNR_?sens\s*=\s*(-?[\d.]+)", cleaned, cast=float),
        "Target": (f"{tgt.group(1)}={tgt.group(2)}" if tgt else None),
    }


LEGEND_COLS = [
    "Test #", "Format", "SU/MU", "FrameBW", "ChannelBW", "Antenna", "Nss",
    "GI [us]", "LTF", "LDPC", "LDPC2xCW", "Length [B]", "PPM", "Channel",
    "Control", "RandomControl", "LFD", "Lean2Main", "RUsize", "RUindex", "SC",
    "SNR_sens [dB]", "Target",
]


def _test_num(name: str) -> float:
    m = re.search(r"[Tt]est\D*(\d+)", name)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d+)", name)
    return int(m.group(1)) if m else 1_000_000


def _metric(path: str) -> str:
    return os.path.basename(path).split(".", 1)[0] or "Y"


def _sheet_name(path: str) -> str:
    base = re.sub(r"\.fig$", "", os.path.basename(path), flags=re.IGNORECASE)
    base = base.replace("WiFi_EHT_Extended_HVT.", "").replace("WiFi_EHT_Extended_HVT", "")
    return base.strip(". ")[:27] or "Sheet"


def _common_grid(curves):
    lo = min(c["x"].min() for c in curves)
    hi = max(c["x"].max() for c in curves)
    lo = np.floor(lo / GRID_STEP) * GRID_STEP
    hi = np.ceil(hi / GRID_STEP) * GRID_STEP
    return np.round(np.arange(lo, hi + GRID_STEP / 2, GRID_STEP), 3)


def write_workbook(out: str, files: list[str]):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    bold = Font(bold=True)

    for path in files:
        curves = extract_curves(path)
        curves.sort(key=lambda c: _test_num(c["name"]))
        metric = _metric(path)
        sname = _sheet_name(path)
        is_err = metric.upper() in ERROR_METRICS

        ws = wb.create_sheet(sname)
        if is_err:
            grid = _common_grid(curves)
            ws.cell(row=1, column=1, value="SNR (dB)").font = bold
            for i, g in enumerate(grid, start=2):
                ws.cell(row=i, column=1, value=float(g))
            for col, c in enumerate(curves, start=2):
                order = np.argsort(c["x"])
                # left = 1.0 (100% error below range), right = 0.0 (0% above range)
                yg = np.interp(grid, c["x"][order], c["y"][order], left=1.0, right=0.0)
                ws.cell(row=1, column=col,
                        value=f"Test {int(_test_num(c['name']))} | {metric}").font = bold
                for i, v in enumerate(yg, start=2):
                    ws.cell(row=i, column=col, value=float(v))
                ws.column_dimensions[get_column_letter(col)].width = 13
            ws.column_dimensions["A"].width = 10
        else:
            col = 1
            for c in curves:
                tag = f"Test {int(_test_num(c['name']))}"
                ws.cell(row=1, column=col, value=f"{tag} | SNR (dB)").font = bold
                ws.cell(row=1, column=col + 1, value=f"{tag} | {metric}").font = bold
                for i, (xv, yv) in enumerate(zip(c["x"], c["y"]), start=2):
                    ws.cell(row=i, column=col, value=float(xv))
                    ws.cell(row=i, column=col + 1, value=float(yv))
                col += 2
        ws.freeze_panes = "B2"

        leg = wb.create_sheet(f"{sname} legend"[:31])
        headers = LEGEND_COLS + ["Points", "Config"]
        for j, h in enumerate(headers, start=1):
            leg.cell(row=1, column=j, value=h).font = bold
        for r, c in enumerate(curves, start=2):
            cleaned = clean_name(c["name"])
            cfg = parse_config(cleaned)
            for j, h in enumerate(LEGEND_COLS, start=1):
                leg.cell(row=r, column=j, value=cfg.get(h))
            leg.cell(row=r, column=len(LEGEND_COLS) + 1, value=int(c["y"].size))
            leg.cell(row=r, column=len(LEGEND_COLS) + 2, value=cleaned)
        leg.column_dimensions[get_column_letter(len(headers))].width = 110
        leg.freeze_panes = "B2"

        print(f"{sname}: {len(curves)} curves"
              f"{' (combined 0.5 dB grid)' if is_err else ' (per-curve X)'}")

    wb.save(out)
    print(f"Saved {out}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("out")
    ap.add_argument("figs", nargs="+")
    args = ap.parse_args()
    write_workbook(args.out, args.figs)
